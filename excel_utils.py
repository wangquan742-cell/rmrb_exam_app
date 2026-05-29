from io import BytesIO
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


def _safe_len(value) -> int:
    if value is None:
        return 0
    return len(str(value))


def beautify_worksheet(ws, sheet_name: str):
    """
    美化单个 Excel sheet。
    """
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 标题行
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 全表格式
    for row in ws.iter_rows(min_row=2):
        score_value = None
        for cell in row:
            if ws.cell(row=1, column=cell.column).value == "考公价值分":
                score_value = cell.value

        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            # 高价值文章浅色标记
            if isinstance(score_value, int) and score_value >= 8:
                cell.fill = PatternFill("solid", fgColor="E2F0D9")

            # 链接列设置超链接
            header = ws.cell(row=1, column=cell.column).value
            if header == "链接" or header == "对应链接":
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    # 行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 54

    ws.row_dimensions[1].height = 28

    # 根据列名设置列宽
    preferred_widths = {
        "日期": 12,
        "版面号": 8,
        "版面": 16,
        "标题": 28,
        "链接": 45,
        "对应链接": 45,
        "来源": 14,
        "考公价值分": 12,
        "主题标签": 22,
        "机关思维": 22,
        "可转化面试题型": 22,
        "正文字数": 10,
        "正文预览": 55,
        "一句话概括": 45,
        "申论积累方向": 50,
        "面试题雏形": 55,
        "中英表达积累": 55,
        "复习状态": 12,
        "个人笔记": 30,
        "备注": 32,
        "复盘类型": 18,
        "内容": 55,
        "建议动作": 50,
    }

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        letter = get_column_letter(col_idx)

        if header in preferred_widths:
            ws.column_dimensions[letter].width = preferred_widths[header]
        else:
            max_len = max(_safe_len(ws.cell(row=row, column=col_idx).value) for row in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 35)


def export_beautified_excel(sheets: Dict[str, pd.DataFrame]) -> bytes:
    """
    导出美化后的 Excel。
    """
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            beautify_worksheet(ws, sheet_name)

    return output.getvalue()
